from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog as fd

#Crear auxiliar para el path al archivo
filename = fd.askopenfilename()

index = 0
while filename[len(filename)-1-index] != "/" :
    index +=1
    
path = filename[0:len(filename)-index]

#path = 'C:/Users/'+getpass.getuser()+'/Documents/py4e/ExcelConPython/'

encabezados = [
    ['Site ID','Site UID','Longitude','Latitude','Description','Site Name'],
    ['Site ID','Antenna ID','Longitude','Latitude','Antenna File','Height(m)','Azimuth','Mechanical tilt','Twist','Donor Antenna'],
    ['Site ID','BTS Name','Sector ID','Relay','Sector UID','Technology','Antenna Algorithm','Band Name','Propagation Model','Distance (km)','Radials','Prediction Mode','Interpolation Distance (m)'],
    ['Site ID','Sector ID','Antenna ID','Link Configuration ID','Cable Length (m)','Bin File Name'],
    ['Site ID','Sector ID','Antenna ID','Port Name','Downlink','Uplink'],
    ['Site ID','BTS Name'],
    ['Site ID','Sector ID','PA Power (dBm)'],
    ['Site ID','Sector ID','Carrier Name','NB-IoT In-band Carrier','Cell ID','Cell Name']
]

mapa = {
    "A":1,
    "B":2,
    "C":3,
    "8":"Jumper_AWS",
    "9":"Jumper_700",
    "5":"Jumper_1900",

    "A08":(1,'LTE-2325_15MHz_SB1_1'),
    "B08":(2,'LTE-2325_15MHz_SB1_1'),
    "C08":(3,'LTE-2325_15MHz_SB1_1'),
    "A09":(11,'LTE_9385_15M_SB1_1'),
    "B09":(12,'LTE_9385_15M_SB1_1'),
    "C09":(13,'LTE_9385_15M_SB1_1'),
    "A05":(21,'LTE_1150_10M_SB1_1'),
    "B05":(22,'LTE_1150_10M_SB1_1'),
    "C05":(23,'LTE_1150_10M_SB1_1'),
    
    "Jumper_AWS": 4,
    "Jumper_700": 1,
    "Jumper_1900": 3,

    "AQU4518R23v06.pafx": {
        "banda_alta_1_+45":	(3,0),
        "banda_alta_1_-45":	(3,0),
        "banda_baja_700_+45":	(1,0),
        "banda_baja_700_-45":	(1,0),
        "banda_baja_850_+45":	(2,0),
        "banda_baja_850_-45":	(2,0),
        "banda_alta_2_+45":	(4,0),
        "banda_alta_2_-45":	(4,0)
    } ,
    "AQU4518R19v06.pafx": {
        "banda_alta_1_+45":	(3,0),
        "banda_alta_1_-45":	(3,0),
        "banda_baja_700_+45":	(1,0),
        "banda_baja_700_-45":	(1,0),
        "banda_baja_850_+45":	(2,0),
        "banda_baja_850_-45":	(2,0),
        "banda_alta_2_+45":	(4,0),
        "banda_alta_2_-45":	(4,0)
    } ,
    "800107681.pafx":{
        "banda_baja_700_+45":	(1,0),
        "banda_baja_700_-45":	(1,0),
        "banda_baja_850_+45":	(2,0),
        "banda_baja_850_-45":	(2,0),
        "banda_alta_1_+45":	(3,0),
        "banda_alta_1_-45":	(3,0),
        "banda_alta_2_+45":	(4,0),
        "banda_alta_2_-45":	(4,0)
    } ,
    "ATR4518R6v06.pafx":{
        "banda_baja_+45":	(1,2),
        "banda_baja_-45":	(1,2),
        "banda_alta_1_+45":	(3,0),
        "banda_alta_1_-45":	(3,0),
        "banda_alta_2_+45":	(4,0),
        "banda_alta_2_-45":	(4,0)
    } ,
    "800108651.pafx":{
        "banda_baja_+45": (1,2),
        "banda_baja_-45":   (1,2),
        "banda_alta_1_+45": (3,0),
        "banda_alta_1_-45": (3,0),
        "banda_alta_2_+45": (4,0),
        "banda_alta_2_-45": (4,0)
    } ,
    "742265V02.pafx":{
        "banda_baja_+45": (2,0),
        "banda_baja_-45":  (2,0),
        "banda_alta_+45":   (3,0),
        "banda_alta_-45":   (3,0)
    } ,
    "800107671.pafx":{
        "R1P45":    (1,0),
        "R1M45":    (1,0),
        "R2P45":    (2,0),
        "R2M45":    (2,0),
        "Y1P45":    (3,0),
        "Y1M45":    (3,0),
        "Y2P45":    (4,0),
        "Y2M45":    (4,0)
    } ,
    "2UNPX203.6R2.pafx":{
        "banda_baja_1_+45": (2,0),
        "banda_baja_1_-45": (2,0),
        "banda_baja_2_+45": (2,0),
        "banda_baja_2_-45": (2,0),
        "banda_alta_1_+45": (3,0),
        "banda_alta_1_-45": (3,0),
        "banda_alta_2_+45": (3,0),
        "banda_alta_2_-45": (3,0)
    } ,
    "ADU451807V01.pafx":{
        "Port 1 - +45": (3,0),
        "Port 2 - -45": (3,0)
    } ,
    "ATR4518R25v06.pafx":{
        "banda_baja_1_+45":  (1,0),
        "banda_baja_1_-45":  (1,0),
        "banda_baja_2_+45":  (2,0),
        "banda_baja_2_-45":  (2,0),
        "banda_alta_+45":   (3,4),
        "banda_alta_-45":   (3,4)
    } ,
    "80010691V011.pafx":{
        "banda_baja_+45":   (1,2),
        "banda_baja_-45":   (1,2),
        "banda_alta_1_+45": (3,0),
        "banda_alta_1_-45": (3,0),
        "banda_alta_2_+45": (4,0),
        "banda_alta_2_-45": (4,0)
    } ,
    "ADU4516R6V06_0790_837.pafx":{
        "banda_baja_1_+45": (1,0),
        "banda_baja_1_-45": (1,0),
        "banda_baja_2_+45": (2,0),
        "banda_baja_2_-45": (2,0)
    } ,
    "JCVV-65A-R4.pafx":{
        "banda_baja_700_+45":   (1,0),
        "banda_baja_700_-45":   (1,0),
        "banda_baja_850_+45":   (2,0),
        "banda_baja_850_-45":   (2,0),
        "banda_alta_1_+45":     (3,0),
        "banda_alta_1_-45":     (3,0),
        "banda_alta_2_+45":     (4,0),
        "banda_alta_2_-45":     (4,0)
    } ,
    "80010892V011.pafx":{
        "R1P45":    (1,2),
        "R1M45":    (1,2),
        "Y1P45":    (3,0),
        "Y1M45":    (3,0),
        "Y2P45":    (4,0),
        "Y2M45":    (4,0)
    } ,
    "7422901.pafx":{
        "Port 1":   (2,3)
    } ,
    "742215.pafx":{
        "Port 1 - +45": (3,0),
        "Port 2 - -45": (3,0)
    } ,
    "CELLMAX-D-CPUSE.pafx":{
        "VERTICAL": (2,3)
    } ,
    "80010681.pafx":{
        "Port 1 - +45": (3,4),
        "Port 2 - -45": (3,4)
    }
}


# Variable auxiliar global
indexExport = 1
celdaExport = ""
#Cargar archivo import
wbImport = load_workbook(filename)
wsImport = wbImport.active
maxRow = wsImport.max_row
maxColumn = wsImport.max_column
#Condiciones para saber que banda se necesita crear
#wsCondiciones = wbImport['Hoja2']
#condiciones = [bool(wsCondiciones['B1']),bool(wsCondiciones['B2']),bool(wsCondiciones['B3'])]


#Crear archivo de salida
wb = Workbook()
ws = wb.active

#Hoja Sites
ws1 = wb.create_sheet("Sites")
ws1.append(encabezados[0])

for row in range(2,maxRow+1):
    rango = 'A'+str(row)
    ws1[rango].value = wsImport[rango].value
    rango = 'C'+str(row)
    ws1[rango].value = wsImport[rango].value
    rango = 'D'+str(row)
    ws1[rango].value = wsImport[rango].value
    rango = 'B'+str(row)
    rangoExp = 'F'+str(row)
    ws1[rangoExp].value = wsImport[rango].value

#Hoja Antenas
ws2 = wb.create_sheet("Antennas")
ws2.append(encabezados[1])
indexEmport = 1
for row in range (2,maxRow+1):
    for col in range (1,maxColumn+2):
        
        letter = get_column_letter(col)
        rango = letter + str(row)
        if col < 7 :                     #Hast la columna F se debe de repetir 3 veces en la hoja de salida
            for rowExp in range (1,4):
                if col == 2:
                    rangoExp = letter + str(indexEmport + rowExp)
                    ws2[rangoExp].value = rowExp
                    
                else:
                    rangoExp = letter + str(indexEmport + rowExp)
                    ws2[rangoExp].value = wsImport[rango].value                  
        elif col <= 12:
            if col <= 9:
                 #letter = get_column_letter(row+5)
                 rangoExp = 'G' + str(col+indexEmport-6)
                 ws2[rangoExp].value = wsImport[rango].value

            else:
                #letter = get_column_letter(row+6)
                rangoExp = 'H' + str(col+indexEmport-9)
                ws2[rangoExp].value = wsImport[rango].value        

        elif col == 13:
            indexEmport += 3
            
#Hoja Sectors
ws3 = wb.create_sheet("Sectors")
ws3.append(encabezados[2])
indexExport = 0


for row in range(2,maxRow+1):
    vectorSectores = []
    vectorBanda = []
    vectorPropagacion = []
    condiciones = [wsImport['M'+str(row)].value,wsImport['N'+str(row)].value,wsImport['O'+str(row)].value]
    maxRowExp = 0
    #print(condiciones)
    if condiciones[0] == 'Y':
        maxRowExp += 3
        auxiliarSectores = ['A08','B08','C08']
        vectorSectores.extend(auxiliarSectores)
        auxiliarSectores = ['LTE-2325_15MHz','LTE-2325_15MHz','LTE-2325_15MHz']
        vectorBanda.extend(auxiliarSectores)
        auxiliarSectores = ['NewUM_masked_2100MHz_tuned.pmf','NewUM_masked_2100MHz_tuned.pmf','NewUM_masked_2100MHz_tuned.pmf']
        vectorPropagacion.extend(auxiliarSectores)

    if condiciones[1] == "Y":
        maxRowExp += 3
        auxiliarSectores = ['A09','B09','C09']
        vectorSectores.extend(auxiliarSectores)
        auxiliarSectores = ['LTE_9385_15M','LTE_9385_15M','LTE_9385_15M']
        vectorBanda.extend(auxiliarSectores)
        auxiliarSectores = ['NewUM_masked_700MHz_tuned.pmf','NewUM_masked_700MHz_tuned.pmf','NewUM_masked_700MHz_tuned.pmf']
        vectorPropagacion.extend(auxiliarSectores)

    if condiciones[2] == "Y":
        maxRowExp += 3
        auxiliarSectores = ['A05','B05','C05']
        vectorSectores.extend(auxiliarSectores)
        auxiliarSectores = ['LTE_1150_10M','LTE_1150_10M','LTE_1150_10M']
        vectorBanda.extend(auxiliarSectores)
        auxiliarSectores = ['UM_Masked_1980_tuned.pmf','UM_Masked_1980_tuned.pmf','UM_Masked_1980_tuned.pmf']
        vectorPropagacion.extend(auxiliarSectores)
    
    #print(vectorSectores)
    for colExp in range (1,len(encabezados[2])+1):
        for rowExp in range (2,maxRowExp + 2):
            if colExp == 1:                
                rangoExp = 'A' + str(rowExp + indexExport)
                #print(rangoExp, wsImport['A'+str(row)].value)
                ws3[rangoExp].value = wsImport['A'+str(row)].value 
                celdaExport = ws3[rangoExp].value
                rangoExp = 'B' + str(rowExp + indexExport)
                ws3[rangoExp].value = 'L'+ celdaExport[1:len(celdaExport)] +'M'
                rangoExp = 'D' + str(rowExp + indexExport)
                ws3[rangoExp].value = 'FALSO'
                rangoExp = 'F' + str(rowExp + indexExport)
                ws3[rangoExp].value = 'LTE FDD'
                rangoExp = 'G' + str(rowExp + indexExport)
                ws3[rangoExp].value = 'NONE'
                rangoExp = 'J' + str(rowExp + indexExport)
                ws3[rangoExp].value = 15
                rangoExp = 'K' + str(rowExp + indexExport)
                ws3[rangoExp].value = 60
                rangoExp = 'M' + str(rowExp + indexExport)
                ws3[rangoExp].value = 200 
            
            if colExp == 3:
                rangoExp = 'B' + str(rowExp + indexExport)
                celdaExport = ws3[rangoExp].value
                rangoExp = 'C' + str(rowExp + indexExport)
                #print(rowExp-2,len(auxiliarSectores), auxiliarSectores[0])
                ws3[rangoExp].value = celdaExport + vectorSectores[rowExp-2] 
                rangoExp = 'H' + str(rowExp + indexExport)
                ws3[rangoExp].value =  vectorBanda[rowExp-2] 
                rangoExp = 'I' + str(rowExp + indexExport)
                ws3[rangoExp].value =  vectorPropagacion[rowExp-2] 
   
    indexExport = indexExport + maxRowExp

#Hoja Sector_Antennas
ws4 = wb.create_sheet("Sector_Antennas")
ws4.append(encabezados[3])
maxRowExp = ws3.max_row

for rowExp in range(2,maxRowExp+1):

    for colExp in range(1,len(encabezados[3])):

        letter = get_column_letter(colExp)
        rangoExp = letter + str(rowExp)
        if colExp == 1:
            ws4[rangoExp].value = ws3[rangoExp].value
            rangoExp = 'E' + str(rowExp)
            ws4[rangoExp].value = 3
        if colExp == 2:
            ws4[rangoExp].value = ws3['C' + str(rowExp)].value
        if colExp == 3:
            celdaExport = ws4['B' + str(rowExp)].value
            #print(celdaExport[6],celdaExport[8])
            #print(mapa[celdaExport[6]],mapa[celdaExport[8]])
            ws4[rangoExp].value = mapa[celdaExport[6]]
        if colExp == 4:
            celdaExport = ws4['B' + str(rowExp)].value
            ws4[rangoExp].value = mapa[celdaExport[8]]

#Hoja Sector_Antennas
ws5 = wb.create_sheet("Sector_Antenna_Ports")
ws5.append(encabezados[4])
maxRowExp = ws4.max_row
antenna = ""


for rowExp in range(2,maxRowExp+1): 
    
    celdaExport = ws4['A' + str(rowExp)].value

    for row in range(2,maxRow+1):
         #print(str(ws4['A' + str(rowExp)].value), str(wsImport['A'+ str(row)].value))
         if celdaExport == wsImport['A'+ str(row)].value:
            antenna = wsImport['E'+ str(row)].value
            #print(antenna)

    #print(len(mapa[antenna]))
    for row in range(1,len(mapa[antenna])+1):
        linea = []
        keys = list(mapa[antenna].keys())
        linea.append(ws4['A' + str(rowExp)].value)
        linea.append(ws4['B' + str(rowExp)].value)
        sector = ws4['B' + str(rowExp)].value
        linea.append(ws4['C' + str(rowExp)].value)
        #print(celdaExport,antenna, keys[row-1])
        linea.append(keys[row-1])
        cond1 = mapa[antenna][keys[row-1]]
        #print(cond1[0])
        #cond2 = celdaExport[8]
        cond2 = mapa[(mapa[sector[8]])]
        #print(cond1, cond2)
        if cond1[0] == cond2 or cond1[1] == cond2 :
            linea.append('VERDADERO')
            linea.append('VERDADERO')
        else:
            linea.append('FALSO')
            linea.append('FALSO')

        ws5.append(linea)

#Hoja LTE_FDD_Base_Stations
ws6 = wb.create_sheet("LTE_FDD_Base_Stations")
ws6.append(encabezados[5])

for row in range(2,maxRow+1):
    linea = []
    celdaExport = wsImport['A' + str(row)].value
    linea.append(celdaExport)
    linea.append('L'+celdaExport[1:len(celdaExport)]+'M')
    ws6.append(linea)



#Hoja LTE_FDD_Sectors
ws7 = wb.create_sheet("LTE_FDD_Sectors")
ws7.append(encabezados[6])
maxRowImport = ws3.max_row

for row in range(2,maxRowImport+1):
    linea = []
    celdaExport = ws3['A' + str(row)].value
    linea.append(celdaExport)
    celdaExport = ws3['C' + str(row)].value
    linea.append(celdaExport)
    linea.append(46)
    ws7.append(linea)

#Hoja LTE_FDD_Sector_Carriers

ws8 = wb.create_sheet("LTE_FDD_Sector_Carriers")
ws8.append(encabezados[7])
maxRowImport = ws7.max_row
for row in range(2,maxRowImport+1):
    linea = []
    celdaExport = ws7['B' + str(row)].value
    linea.append(ws7['A' + str(row)].value)
    linea.append(ws7['B' + str(row)].value)
    linea.append("")
    vector = mapa[celdaExport[6:len(celdaExport)]]
    linea.append(vector[1])
    linea.append(vector[0])
    linea.append(celdaExport)
    ws8.append(linea)

wb.save(path+'exportSheet.xlsx')